"""
Generate Supplementary Table S3: Comprehensive Performance Metrics
This script creates a publication-ready supplementary table with all evaluation metrics
including precision, recall, F1-score, and inter-rater reliability measures.
"""

import os
import pandas as pd
import numpy as np
import sys
sys.path.append('/Users/fernando/Documents/Research/academatepy')
from validation.analysis_functions import (
    calculate_performance_metrics,
    generate_uniqueid,
    model_name_corrections,
    process_review_repredicting
)
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Configuration - Models to include (matching paper's Figure 3)
MODEL_LIST = [
    "gpt-3.5-turbo-0125",
    "gpt-4o-2024-11-20",
    "gpt-4o-mini-2024-07-18",
    "gemini-2.0-flash",
    "gemini-2.0-flash-lite",
    "gemini-1.5-pro",
]

# Convert to folder names
modelsname_infolder = [x.replace("/", "_").replace(":", "_") for x in MODEL_LIST]
folder2modelname = dict(zip(modelsname_infolder, MODEL_LIST))

# Working directories
workdir = "/Users/fernando/Documents/Research/academatepy/validation/"
result_type = "results"

# Review configurations
rw_1_workdir_scr1 = {
    'directory': f"{workdir}PICOS/",
    'results_directory': f"{workdir}PICOS/{result_type}/",
    "columns_needed_as_true": ['population_scr1', 'intervention_scr1', 'physio_and_other_scr1',
                               'e_interventions_scr1', 'control_group_scr1', 'outcome_scr1',
                               'study_type_scr1'],
    "model_list": modelsname_infolder,
    "description": "Physiotherapy",
    "name": "I"
}

rw_1_workdir_scr2 = {
    'directory': f"{workdir}PICOS/",
    'results_directory': f"{workdir}PICOS/{result_type}/",
    "columns_needed_as_true": ['population_scr2', 'intervention_scr2', 'physio_and_other_scr2',
                               'e_interventions_scr2', 'control_group_scr2', 'outcome_scr2',
                               'study_type_scr2'],
    "model_list": modelsname_infolder,
    "description": "Physiotherapy",
    "name": "I"
}

rw_2_workdir_scr1 = {
    'directory': f"{workdir}reproduction/",
    'results_directory': f"{workdir}reproduction/{result_type}/",
    "columns_needed_as_true": ['Population_scr1', 'Intervention_scr1', 'Human_scr1',
                               'Preclinical_Clinical_scr1', 'Outcome_scr1', 'Publicationtype_scr1'],
    "model_list": modelsname_infolder,
    "description": "Endometrial disorders",
    "name": "II"
}

rw_2_workdir_scr2 = {
    'directory': f"{workdir}reproduction/",
    'results_directory': f"{workdir}reproduction/{result_type}/",
    "columns_needed_as_true": ['Population_scr2', 'Intervention_scr2', 'Human_scr2',
                               'Preclinical_Clinical_scr2', 'Outcome_scr2', 'Publicationtype_scr2'],
    "model_list": modelsname_infolder,
    "description": "Endometrial disorders",
    "name": "II"
}


def process_all_reviews():
    """Process all reviews and return combined performance metrics."""
    all_results = []

    reviews_config = [
        (rw_1_workdir_scr1, 'screening1', 'I', 'Physiotherapy'),
        (rw_1_workdir_scr2, 'screening2', 'I', 'Physiotherapy'),
        (rw_2_workdir_scr1, 'screening1', 'II', 'Endometrial disorders'),
        (rw_2_workdir_scr2, 'screening2', 'II', 'Endometrial disorders'),
    ]

    for review_config, screening_type, review_name, review_desc in reviews_config:
        print(f"\nProcessing Review {review_name} ({review_desc}) - {screening_type}")
        try:
            _, performance_results = process_review_repredicting(review_config, screening_type=screening_type)
            performance_results['review_description'] = review_desc
            all_results.append(performance_results)
            print(f"  Successfully processed {len(performance_results)} models")
        except Exception as e:
            print(f"  Error: {str(e)}")
            continue

    if all_results:
        combined_df = pd.concat(all_results, ignore_index=True)
        return combined_df
    return pd.DataFrame()


def create_summary_tables(df):
    """Create summary tables for the supplementary materials."""

    # Metrics to include in the table
    metrics_detail = ['TP', 'TN', 'FP', 'FN', 'precision', 'recall', 'f1_score',
                      'mcc', 'cohen_kappa', 'analysis_coverage']

    metrics_adjusted = ['adjusted_precision', 'adjusted_recall', 'adjusted_f1_score',
                        'adjusted_mcc', 'adjusted_cohen_kappa']

    # Table 1: Detailed metrics by model, review, and screening type
    detail_cols = ['model_name', 'reviewname', 'review_description', 'screening_type',
                   'total_records', 'analyzed_records', 'missing_records'] + metrics_detail + metrics_adjusted

    # Filter to only include the models we want
    df_filtered = df[df['model_name'].isin(MODEL_LIST)].copy()

    # Create detailed table
    detail_table = df_filtered[detail_cols].copy()
    detail_table = detail_table.sort_values(['model_name', 'reviewname', 'screening_type'])

    # Round numeric columns
    numeric_cols = metrics_detail + metrics_adjusted
    for col in numeric_cols:
        if col in detail_table.columns:
            detail_table[col] = detail_table[col].round(3)

    # Table 2: Summary by model (averaged across reviews and screening types)
    summary_metrics = ['precision', 'recall', 'f1_score', 'mcc', 'cohen_kappa',
                       'adjusted_precision', 'adjusted_recall', 'adjusted_f1_score',
                       'adjusted_mcc', 'adjusted_cohen_kappa']

    summary_table = df_filtered.groupby('model_name')[summary_metrics].mean().round(3)
    summary_table = summary_table.sort_values('adjusted_mcc', ascending=False)

    # Table 3: Summary by screening type
    screening_summary = df_filtered.groupby(['screening_type', 'model_name'])[summary_metrics].mean().round(3)
    screening_summary = screening_summary.reset_index()

    return detail_table, summary_table, screening_summary


def create_styled_excel(detail_table, summary_table, screening_summary, output_path):
    """Create a publication-ready styled Excel file."""

    wb = Workbook()

    # Define styles matching paper aesthetics
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill_blue = PatternFill(start_color='5A69A4', end_color='5A69A4', fill_type='solid')
    header_fill_orange = PatternFill(start_color='E9A064', end_color='E9A064', fill_type='solid')
    header_fill_green = PatternFill(start_color='4A7C59', end_color='4A7C59', fill_type='solid')

    cell_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Alternating row colors
    fill_light = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # === Sheet 1: Detailed Results ===
    ws1 = wb.active
    ws1.title = "S3a_Detailed_Metrics"

    # Rename columns for publication
    column_rename = {
        'model_name': 'Model',
        'reviewname': 'Review',
        'review_description': 'Review Topic',
        'screening_type': 'Screening Phase',
        'total_records': 'Total Records',
        'analyzed_records': 'Analyzed Records',
        'missing_records': 'Failed Records',
        'TP': 'True Positives',
        'TN': 'True Negatives',
        'FP': 'False Positives',
        'FN': 'False Negatives',
        'precision': 'Precision',
        'recall': 'Recall (Sensitivity)',
        'f1_score': 'F1-Score',
        'mcc': 'MCC',
        'cohen_kappa': "Cohen's Kappa",
        'analysis_coverage': 'Analysis Coverage',
        'adjusted_precision': 'Adj. Precision',
        'adjusted_recall': 'Adj. Recall',
        'adjusted_f1_score': 'Adj. F1-Score',
        'adjusted_mcc': 'Adj. MCC',
        'adjusted_cohen_kappa': "Adj. Cohen's Kappa"
    }

    detail_renamed = detail_table.rename(columns=column_rename)

    # Write headers
    for col_idx, col_name in enumerate(detail_renamed.columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = center_align
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(detail_renamed.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx > 1 else left_align
            cell.fill = fill_light if row_idx % 2 == 0 else fill_white

    # Adjust column widths
    for col_idx in range(1, len(detail_renamed.columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = 15
    ws1.column_dimensions['A'].width = 25  # Model name
    ws1.column_dimensions['C'].width = 20  # Review topic

    # === Sheet 2: Summary by Model ===
    ws2 = wb.create_sheet("S3b_Model_Summary")

    summary_renamed = summary_table.rename(columns={
        'precision': 'Mean Precision',
        'recall': 'Mean Recall',
        'f1_score': 'Mean F1-Score',
        'mcc': 'Mean MCC',
        'cohen_kappa': "Mean Cohen's Kappa",
        'adjusted_precision': 'Mean Adj. Precision',
        'adjusted_recall': 'Mean Adj. Recall',
        'adjusted_f1_score': 'Mean Adj. F1-Score',
        'adjusted_mcc': 'Mean Adj. MCC',
        'adjusted_cohen_kappa': "Mean Adj. Cohen's Kappa"
    })

    # Write headers
    ws2.cell(row=1, column=1, value='Model').font = header_font
    ws2.cell(row=1, column=1).fill = header_fill_orange
    ws2.cell(row=1, column=1).alignment = center_align
    ws2.cell(row=1, column=1).border = thin_border

    for col_idx, col_name in enumerate(summary_renamed.columns, 2):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill_orange
        cell.alignment = center_align
        cell.border = thin_border

    # Write data
    for row_idx, (model_name, row) in enumerate(summary_renamed.iterrows(), 2):
        ws2.cell(row=row_idx, column=1, value=model_name).font = cell_font
        ws2.cell(row=row_idx, column=1).border = thin_border
        ws2.cell(row=row_idx, column=1).fill = fill_light if row_idx % 2 == 0 else fill_white

        for col_idx, value in enumerate(row.values, 2):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = fill_light if row_idx % 2 == 0 else fill_white

    # Adjust column widths
    ws2.column_dimensions['A'].width = 25
    for col_idx in range(2, len(summary_renamed.columns) + 2):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    # === Sheet 3: Summary by Screening Type ===
    ws3 = wb.create_sheet("S3c_Screening_Summary")

    screening_renamed = screening_summary.rename(columns={
        'screening_type': 'Screening Phase',
        'model_name': 'Model',
        'precision': 'Mean Precision',
        'recall': 'Mean Recall',
        'f1_score': 'Mean F1-Score',
        'mcc': 'Mean MCC',
        'cohen_kappa': "Mean Cohen's Kappa",
        'adjusted_precision': 'Mean Adj. Precision',
        'adjusted_recall': 'Mean Adj. Recall',
        'adjusted_f1_score': 'Mean Adj. F1-Score',
        'adjusted_mcc': 'Mean Adj. MCC',
        'adjusted_cohen_kappa': "Mean Adj. Cohen's Kappa"
    })

    # Write headers
    for col_idx, col_name in enumerate(screening_renamed.columns, 1):
        cell = ws3.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill_green
        cell.alignment = center_align
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(screening_renamed.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx > 2 else left_align
            cell.fill = fill_light if row_idx % 2 == 0 else fill_white

    # Adjust column widths
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 25
    for col_idx in range(3, len(screening_renamed.columns) + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 18

    # === Sheet 4: Legend ===
    ws4 = wb.create_sheet("Legend")

    legend_data = [
        ("Metric", "Description", "Interpretation"),
        ("True Positives (TP)", "Articles correctly classified as included", "Higher is better"),
        ("True Negatives (TN)", "Articles correctly classified as excluded", "Higher is better"),
        ("False Positives (FP)", "Articles incorrectly classified as included", "Lower is better"),
        ("False Negatives (FN)", "Articles incorrectly classified as excluded (missed)", "Lower is better - critical for systematic reviews"),
        ("Precision", "TP / (TP + FP) - Positive predictive value", "Proportion of predicted inclusions that are correct"),
        ("Recall (Sensitivity)", "TP / (TP + FN) - True positive rate", "Proportion of actual inclusions correctly identified"),
        ("F1-Score", "Harmonic mean of precision and recall", "Balanced measure of precision and recall"),
        ("MCC", "Matthews Correlation Coefficient", "Balanced measure for imbalanced datasets (-1 to 1)"),
        ("Cohen's Kappa", "Inter-rater reliability accounting for chance", "≤0: no agreement, 0.01-0.20: slight, 0.21-0.40: fair, 0.41-0.60: moderate, 0.61-0.80: substantial, 0.81-1.00: almost perfect"),
        ("Adjusted Metrics", "Original metric × (analyzed records / total records)", "Penalizes models for failing to process records"),
        ("Analysis Coverage", "Proportion of records successfully analyzed", "Higher is better (1.0 = all records processed)"),
    ]

    for row_idx, row_data in enumerate(legend_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill_blue
            else:
                cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 45
    ws4.column_dimensions['C'].width = 50

    # Set row heights for legend
    for row in range(1, len(legend_data) + 1):
        ws4.row_dimensions[row].height = 30

    # Save workbook
    wb.save(output_path)
    print(f"\nSupplementary table saved to: {output_path}")


def generate_latex_table(summary_table, output_path):
    """Generate a LaTeX table for direct inclusion in the paper."""

    # Select key metrics
    metrics_for_latex = ['precision', 'recall', 'f1_score', 'adjusted_mcc', 'cohen_kappa']
    latex_df = summary_table[metrics_for_latex].copy()

    # Rename columns
    latex_df.columns = ['Precision', 'Recall', 'F1', 'Adj. MCC', "Cohen's $\\kappa$"]

    # Generate LaTeX
    latex_str = latex_df.to_latex(
        float_format="%.3f",
        column_format='l' + 'c' * len(latex_df.columns),
        caption="Summary of model performance metrics averaged across all reviews and screening phases.",
        label="tab:metrics_summary",
        escape=False
    )

    with open(output_path, 'w') as f:
        f.write(latex_str)

    print(f"LaTeX table saved to: {output_path}")


def print_summary_statistics(df):
    """Print summary statistics for the paper text."""

    df_filtered = df[df['model_name'].isin(MODEL_LIST)].copy()

    print("\n" + "="*80)
    print("SUMMARY STATISTICS FOR PAPER TEXT")
    print("="*80)

    # Overall summary by model
    print("\n--- Mean Performance by Model (across all reviews and screening types) ---")
    summary = df_filtered.groupby('model_name').agg({
        'precision': 'mean',
        'recall': 'mean',
        'f1_score': 'mean',
        'adjusted_mcc': 'mean',
        'cohen_kappa': 'mean',
        'adjusted_precision': 'mean',
        'adjusted_recall': 'mean'
    }).round(3)
    summary = summary.sort_values('adjusted_mcc', ascending=False)
    print(summary.to_string())

    # Summary by screening type
    print("\n--- Mean Performance by Screening Type ---")
    screening_summary = df_filtered.groupby('screening_type').agg({
        'precision': 'mean',
        'recall': 'mean',
        'f1_score': 'mean',
        'adjusted_mcc': 'mean',
        'cohen_kappa': 'mean'
    }).round(3)
    print(screening_summary.to_string())

    # Best performing model text
    best_model = summary['adjusted_mcc'].idxmax()
    best_metrics = summary.loc[best_model]

    print(f"\n--- Text for Results Section ---")
    print(f"The best-performing model, {best_model}, achieved a mean precision of {best_metrics['precision']:.3f}, ")
    print(f"mean recall of {best_metrics['recall']:.3f}, and mean F1-score of {best_metrics['f1_score']:.3f}.")
    print(f"When adjusted for analysis coverage, {best_model} achieved an adjusted precision of {best_metrics['adjusted_precision']:.3f} ")
    print(f"and adjusted recall of {best_metrics['adjusted_recall']:.3f}.")


if __name__ == "__main__":
    # Create output directory
    output_dir = f"{workdir}results/{result_type}"
    os.makedirs(output_dir, exist_ok=True)

    print("Processing reviews to generate supplementary metrics table...")
    print("="*60)

    # Process all reviews
    combined_results = process_all_reviews()

    if combined_results.empty:
        print("ERROR: No results to process!")
        sys.exit(1)

    # Create summary tables
    detail_table, summary_table, screening_summary = create_summary_tables(combined_results)

    # Save raw data
    combined_results.to_pickle(f"{output_dir}/all_performance_metrics.pkl")
    combined_results.to_csv(f"{output_dir}/all_performance_metrics.csv", index=False)

    # Create styled Excel
    excel_path = f"{output_dir}/Supplementary_Table_S3_Performance_Metrics.xlsx"
    create_styled_excel(detail_table, summary_table, screening_summary, excel_path)

    # Generate LaTeX table
    latex_path = f"{output_dir}/table_s3_latex.tex"
    generate_latex_table(summary_table, latex_path)

    # Print summary statistics
    print_summary_statistics(combined_results)

    print("\n" + "="*60)
    print("DONE! Files generated:")
    print(f"  1. {excel_path}")
    print(f"  2. {latex_path}")
    print(f"  3. {output_dir}/all_performance_metrics.csv")
    print("="*60)
