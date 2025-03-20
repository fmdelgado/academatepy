import pandas as pd
import os
import sys
from tqdm import tqdm
import time

# Add the project root to the Python path
sys.path.insert(0, '/Users/fernando/Documents/Research/academatepy')

# Import the ArticleMetadataRetriever
from academate.research_module.metadata_retriever import ArticleMetadataRetriever

# Ruta al archivo Excel
excel_file = "/Users/fernando/Documents/Research/academatepy/validation/reproduction/rodriguez_2024.xlsx"

# Verificar si el archivo existe
if not os.path.exists(excel_file):
    print(f"Error: El archivo {excel_file} no existe.")
    exit(1)

# Leer el archivo Excel
try:
    df = pd.read_excel(excel_file, engine='openpyxl')
    print(f"Archivo cargado correctamente. Dimensiones: {df.shape}")
except Exception as e:
    print(f"Error al leer el archivo Excel: {e}")
    exit(1)

df.rename(columns={'TÍTULO': 'title', 'SCREENING 1': 'screening1', 'SCREENING 2': 'screening2',
                   "Guardado en carpeta": "pdf_name"}, inplace=True)
# Analizar los datos
print(df.columns)

total_records = len(df)

print(f"\nEstadísticas de la base de datos:")
print(f"Total de registros: {total_records}")

# Verificar si existen las columnas de screening
if 'screening1' in df.columns:
    screening1_passed = df[df['screening1'] == True].shape[0]
    screening1_failed = df[df['screening1'] == False].shape[0]
    screening1_na = df[df['screening1'].isna()].shape[0]

    print(f"\nScreening 1:")
    print(f"- Pasaron: {screening1_passed} ({screening1_passed / total_records * 100:.1f}%)")
    print(f"- No pasaron: {screening1_failed} ({screening1_failed / total_records * 100:.1f}%)")
    print(f"- Sin evaluar: {screening1_na} ({screening1_na / total_records * 100:.1f}%)")
else:
    print("\nNo se encontró la columna 'screening1' en el dataset")

if 'screening2' in df.columns:
    screening2_passed = df[df['screening2'] == True].shape[0]
    screening2_failed = df[df['screening2'] == False].shape[0]
    screening2_na = df[df['screening2'].isna()].shape[0]

    print(f"\nScreening 2:")
    print(f"- Pasaron: {screening2_passed} ({screening2_passed / total_records * 100:.1f}%)")
    print(f"- No pasaron: {screening2_failed} ({screening2_failed / total_records * 100:.1f}%)")
    print(f"- Sin evaluar: {screening2_na} ({screening2_na / total_records * 100:.1f}%)")

    # Análisis de los que pasaron ambos screenings
    if 'screening1' in df.columns:
        both_passed = df[(df['screening1'] == True) & (df['screening2'] == True)].shape[0]
        print(f"\nRegistros que pasaron ambos screenings: {both_passed} ({both_passed / total_records * 100:.1f}%)")
else:
    print("\nNo se encontró la columna 'screening2' en el dataset")

# Verificar si hay columnas adicionales de interés
print("\nOtras columnas de interés:")
for col in ['doi', 'pmid', 'pdf_path', 'pdf_name']:
    if col in df.columns:
        non_null_count = df[col].notna().sum()
        print(f"- {col}: {non_null_count} valores no nulos ({non_null_count / total_records * 100:.1f}%)")

# Fetch metadata using ArticleMetadataRetriever
# First, try to get metadata using PMIDs
print("\nExtrayendo metadatos (abstracts y otra información) desde PubMed y otras fuentes...")

# Fix PMID format - convert from float to integer strings
if 'PMID' in df.columns:
    # Convert float PMIDs to integer strings (remove decimal part)
    df['PMID'] = df['PMID'].apply(lambda x: str(int(x)) if pd.notna(x) else x)
    print("PMIDs convertidos de formato decimal a entero")

# Initialize the metadata retriever with your email
email = "your_email@example.com"  # Replace with your email
retriever = ArticleMetadataRetriever(email=email, verbose=True)

# Initialize new columns for metadata if they don't exist
if 'abstract' not in df.columns:
    df['abstract'] = None
if 'journal' not in df.columns:
    df['journal'] = None
if 'publication_date' not in df.columns:
    df['publication_date'] = None
if 'metadata_source' not in df.columns:
    df['metadata_source'] = None

# Create a checkpoint file path
checkpoint_file = "/Users/fernando/Documents/Research/academatepy/validation/reproduction/metadata_checkpoint.pkl"

# Load previous progress if exists
if os.path.exists(checkpoint_file):
    try:
        checkpoint_df = pd.read_pickle(checkpoint_file)
        # Update only the metadata columns from the checkpoint
        for col in ['abstract', 'journal', 'publication_date', 'metadata_source']:
            if col in checkpoint_df.columns:
                # Use index matching to update values
                for idx in checkpoint_df.index:
                    if idx in df.index and pd.notna(checkpoint_df.loc[idx, col]):
                        df.loc[idx, col] = checkpoint_df.loc[idx, col]

        print(f"Cargado progreso anterior desde {checkpoint_file}")
        print(f"- Registros con abstract: {df['abstract'].notna().sum()}")
    except Exception as e:
        print(f"Error al cargar el checkpoint: {e}")
        print("Continuando sin checkpoint...")

# First, try to get metadata using PMIDs - only for records without abstracts
pmid_mask = df['PMID'].notna() & df['abstract'].isna()
pmid_count = pmid_mask.sum()
print(f"Encontrados {pmid_count} registros con PMID pendientes de procesar")

if pmid_count > 0:
    print("Obteniendo metadatos usando PMIDs...")

    # Process records with PMIDs
    for idx, row in tqdm(df[pmid_mask].iterrows(), total=pmid_count, desc="Procesando PMIDs"):
        pmid = str(row['PMID']).strip()

        try:
            # Use Entrez directly for PMID lookup
            from Bio import Entrez

            Entrez.email = email

            # Fetch the record
            handle = Entrez.efetch(db="pubmed", id=pmid, rettype="xml", retmode="text")
            record = Entrez.read(handle)
            handle.close()

            if record and 'PubmedArticle' in record and record['PubmedArticle']:
                article = record['PubmedArticle'][0]

                # Extract article data
                if 'MedlineCitation' in article and 'Article' in article['MedlineCitation']:
                    article_data = article['MedlineCitation']['Article']

                    # Get abstract - handle case when AbstractText might be missing
                    if 'Abstract' in article_data:
                        if 'AbstractText' in article_data['Abstract']:
                            abstract_sections = article_data['Abstract']['AbstractText']
                            abstract_text = []

                            for section in abstract_sections:
                                # Handle labeled sections
                                if hasattr(section, 'attributes') and 'Label' in section.attributes:
                                    label = section.attributes['Label']
                                    abstract_text.append(f"{label}: {section}")
                                else:
                                    abstract_text.append(str(section))

                            df.at[idx, 'abstract'] = ' '.join(abstract_text)

                    # Get journal
                    if 'Journal' in article_data and 'Title' in article_data['Journal']:
                        df.at[idx, 'journal'] = article_data['Journal']['Title']

                    # Get publication date
                    if 'Journal' in article_data and 'JournalIssue' in article_data['Journal'] and 'PubDate' in \
                            article_data['Journal']['JournalIssue']:
                        pub_date = article_data['Journal']['JournalIssue']['PubDate']
                        date_parts = []

                        if 'Year' in pub_date:
                            date_parts.append(pub_date['Year'])
                        if 'Month' in pub_date:
                            date_parts.append(pub_date['Month'])
                        if 'Day' in pub_date:
                            date_parts.append(pub_date['Day'])

                        df.at[idx, 'publication_date'] = ' '.join(date_parts)

                    df.at[idx, 'metadata_source'] = 'pubmed_pmid'

                    # Save checkpoint periodically (every 10 records)
                    if idx % 10 == 0:
                        df.to_pickle(checkpoint_file)
        except Exception as e:
            print(f"Error al obtener metadatos para PMID {pmid}: {str(e)}")
            # Add a small delay to avoid overwhelming the API
            import time

            time.sleep(0.5)

    # Save checkpoint after processing all PMIDs
    df.to_pickle(checkpoint_file)
    print(f"Progreso guardado en {checkpoint_file}")

# Then try DOIs for records that don't have metadata yet
doi_mask = df['DOI'].notna() & df['abstract'].isna()
doi_count = doi_mask.sum()
print(f"Encontrados {doi_count} registros adicionales con DOI pero sin abstract")

if doi_count > 0:
    print("Obteniendo metadatos usando DOIs...")

    for idx, row in tqdm(df[doi_mask].iterrows(), total=doi_count, desc="Procesando DOIs"):
        doi = str(row['DOI']).strip()

        try:
            # Get metadata using the retriever
            metadata = retriever.get_best_metadata(doi)

            if metadata.get('abstract'):
                df.at[idx, 'abstract'] = metadata['abstract']
            if metadata.get('journal'):
                df.at[idx, 'journal'] = metadata['journal']
            if metadata.get('publication_date'):
                df.at[idx, 'publication_date'] = metadata['publication_date']

            df.at[idx, 'metadata_source'] = ','.join(metadata.get('metadata_sources', []))

            # Save checkpoint periodically (every 10 records)
            if idx % 10 == 0:
                df.to_pickle(checkpoint_file)
        except Exception as e:
            print(f"Error al obtener metadatos para DOI {doi}: {str(e)[:100]}...")

    # Save checkpoint after processing all DOIs
    df.to_pickle(checkpoint_file)
    print(f"Progreso guardado en {checkpoint_file}")

# Print summary of metadata retrieval
abstract_count = df['abstract'].notna().sum()
abstract_missing = df['abstract'].isna().sum()
print(f"\nResumen de extracción de metadatos:")
print(f"- Registros con abstract: {abstract_count} ({abstract_count / total_records * 100:.1f}%)")
print(f"- Registros sin abstract: {abstract_missing} ({abstract_missing / total_records * 100:.1f}%)")
print(f"- Registros con journal: {df['journal'].notna().sum()} ({df['journal'].notna().sum() / total_records * 100:.1f}%)")
print(f"- Registros con fecha de publicación: {df['publication_date'].notna().sum()} ({df['publication_date'].notna().sum() / total_records * 100:.1f}%)")

# Count abstracts in screening subgroups
if 'screening1' in df.columns:
    screening1_passed_count = df[df['screening1'] == True].shape[0]
    screening1_passed_with_abstract = df[(df['screening1'] == True) & (df['abstract'].notna())].shape[0]
    screening1_passed_without_abstract = df[(df['screening1'] == True) & (df['abstract'].isna())].shape[0]
    
    print(f"\nAbstracts en artículos que pasaron Screening 1:")
    print(f"- Con abstract: {screening1_passed_with_abstract} ({screening1_passed_with_abstract / screening1_passed_count * 100:.1f}% de los que pasaron)")
    print(f"- Sin abstract: {screening1_passed_without_abstract} ({screening1_passed_without_abstract / screening1_passed_count * 100:.1f}% de los que pasaron)")

if 'screening2' in df.columns:
    screening2_passed_count = df[df['screening2'] == True].shape[0]
    screening2_passed_with_abstract = df[(df['screening2'] == True) & (df['abstract'].notna())].shape[0]
    screening2_passed_without_abstract = df[(df['screening2'] == True) & (df['abstract'].isna())].shape[0]
    
    print(f"\nAbstracts en artículos que pasaron Screening 2:")
    print(f"- Con abstract: {screening2_passed_with_abstract} ({screening2_passed_with_abstract / screening2_passed_count * 100:.1f}% de los que pasaron)")
    print(f"- Sin abstract: {screening2_passed_without_abstract} ({screening2_passed_without_abstract / screening2_passed_count * 100:.1f}% de los que pasaron)")

if 'screening1' in df.columns and 'screening2' in df.columns:
    both_passed_count = df[(df['screening1'] == True) & (df['screening2'] == True)].shape[0]
    both_passed_with_abstract = df[(df['screening1'] == True) & (df['screening2'] == True) & (df['abstract'].notna())].shape[0]
    both_passed_without_abstract = df[(df['screening1'] == True) & (df['screening2'] == True) & (df['abstract'].isna())].shape[0]
    
    print(f"\nAbstracts en artículos que pasaron ambos screenings:")
    print(f"- Con abstract: {both_passed_with_abstract} ({both_passed_with_abstract / both_passed_count * 100:.1f}% de los que pasaron ambos)")
    print(f"- Sin abstract: {both_passed_without_abstract} ({both_passed_without_abstract / both_passed_count * 100:.1f}% de los que pasaron ambos)")

# Create a new column 'Record' that combines title and abstract
print("\nCreando columna 'Record' combinando título y abstract...")
df['Record'] = df.apply(
    lambda row: f"{row['title']}\n\n{row['abstract']}" if pd.notna(row['title']) and pd.notna(row['abstract']) else None, 
    axis=1
)

# Export records with non-empty title AND abstract to pickle
records_with_title_and_abstract = df[df['title'].notna() & df['abstract'].notna()]
records_count = len(records_with_title_and_abstract)
print(f"\nExportando {records_count} registros con título y abstract no vacíos ({records_count / total_records * 100:.1f}% del total)")

# Save to pickle
pickle_file = "/Users/fernando/Documents/Research/academatepy/validation/reproduction/records_with_abstracts.pkl"
records_with_title_and_abstract.to_pickle(pickle_file)
print(f"Registros guardados en {pickle_file}")

# Create a summary of metadata sources
if 'metadata_source' in df.columns:
    source_counts = df['metadata_source'].value_counts()
    print("\nFuentes de metadatos:")
    for source, count in source_counts.items():
        if pd.notna(source):
            print(f"- {source}: {count} registros")

# Guardar un resumen en un archivo de texto
with open("/Users/fernando/Documents/Research/academatepy/database_summary.txt", "w") as f:
    f.write(f"Resumen de Database Rodríguez-Eguren 2024.xlsx\n")
    f.write(f"==============================================\n\n")
    f.write(f"Total de registros: {total_records}\n\n")

    if 'screening1' in df.columns:
        f.write(f"Screening 1:\n")
        f.write(f"- Pasaron: {screening1_passed} ({screening1_passed / total_records * 100:.1f}%)\n")
        f.write(f"- No pasaron: {screening1_failed} ({screening1_failed / total_records * 100:.1f}%)\n")
        f.write(f"- Sin evaluar: {screening1_na} ({screening1_na / total_records * 100:.1f}%)\n\n")

    if 'screening2' in df.columns:
        f.write(f"Screening 2:\n")
        f.write(f"- Pasaron: {screening2_passed} ({screening2_passed / total_records * 100:.1f}%)\n")
        f.write(f"- No pasaron: {screening2_failed} ({screening2_failed / total_records * 100:.1f}%)\n")
        f.write(f"- Sin evaluar: {screening2_na} ({screening2_na / total_records * 100:.1f}%)\n\n")

        if 'screening1' in df.columns:
            f.write(
                f"Registros que pasaron ambos screenings: {both_passed} ({both_passed / total_records * 100:.1f}%)\n\n")

    # Add metadata summary to the text file
    f.write(f"Metadatos:\n")
    f.write(f"- Registros con abstract: {abstract_count} ({abstract_count / total_records * 100:.1f}%)\n")
    f.write(
        f"- Registros con journal: {df['journal'].notna().sum()} ({df['journal'].notna().sum() / total_records * 100:.1f}%)\n")
    f.write(
        f"- Registros con fecha de publicación: {df['publication_date'].notna().sum()} ({df['publication_date'].notna().sum() / total_records * 100:.1f}%)\n\n")


df.abstract.isna().value_counts()
df[df.screening1 == True].abstract.isna().value_counts()
df[df.screening2 == True].abstract.isna().value_counts()

# Create a new workbook and add a worksheet
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Articles"

# Add the DataFrame to the worksheet
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
    ws.append(row)
    if r_idx == 0:  # Apply formatting to header row
        for cell in ws[r_idx + 1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter

    # Find the maximum content length in the column
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    # Set width based on content, with limits
    adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, Max 50
    ws.column_dimensions[column_letter].width = adjusted_width

# Format abstract column to wrap text
abstract_col = None
for idx, col_name in enumerate(df.columns):
    if col_name == 'abstract':
        abstract_col = idx + 1  # +1 because Excel columns are 1-indexed
        break

if abstract_col:
    for row in ws.iter_rows(min_row=2, min_col=abstract_col, max_col=abstract_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Make abstract column wider
    ws.column_dimensions[ws.cell(row=1, column=abstract_col).column_letter].width = 60

# Save the workbook
wb.save("/Users/fernando/Documents/Research/academatepy/database_summary.xlsx")
print("\nArchivo Excel guardado con metadatos adicionales.")